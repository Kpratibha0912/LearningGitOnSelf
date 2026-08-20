import openpyxl

book = openpyxl.load_workbook("P:\\SeleniumWithPython\\TestData_excel_PythonLearning.xlsx")
sheet = book.active
cellValue = sheet.cell(row=1, column=1).value
# print(cellValue)

# adding data to the excel
sheet.cell(row = 5, column = 4).value= "Adding data from automation"
# print(sheet.cell(row = 5, column = 4).value)

sheetDict = {}

# Print all the data from the excel
for i in range(1, sheet.max_row + 1):

    # To fetch the data only for Sheet2_T_02
    if sheet.cell(row=i, column = 1).value == "Sheet2_T_02":
        for j in range(2, sheet.max_column + 1):
            sheetDict[sheet.cell(row = 1, column = j).value] = sheet.cell(row = i, column = j).value
            # print(sheet.cell(row = i, column = j).value)

print(sheetDict)


